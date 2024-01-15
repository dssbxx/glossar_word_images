import json

from openpyxl import load_workbook
import os


jieshi = {}
with open("./jieshi.txt", 'r') as fin:
    for line in fin:
        segs = [sp.strip() for sp in line.split("\t")]
        assert len(segs) == 2
        keyword, exp = segs[0], segs[1]
        jieshi[keyword] = exp

# 文件路径
xlsx_files = ["./p1.xlsx", "./p2.xlsx", "./p3.xlsx", "./p4.xlsx"]

image_dir = 'extracted_images'

img_idx = 0
fout = open("./record.jsonl", 'w')
for f in xlsx_files:
    wb = load_workbook(filename=f, data_only=True)
    ws = wb.active
    pictures = {(img.anchor._from.col, img.anchor._from.row): img for img in ws._images}


    start_row = 1
    end_row = 3000
    image_column_index = 1  # 第二列的索引，Excel是从1开始计数

    for row in range(start_row, end_row + 1):
        # 如果当前行和列对应的单元格有图像
        if (image_column_index, row) in pictures:
            keyword = ws[row+1][0].value.strip()
            if keyword not in jieshi or len(keyword) < 2:
               continue
            # 获取对应的图片对象
            image = pictures[(image_column_index, row)]
            # 图片的数据在`_data`属性中
            img_data = image._data()
            # 构建每张图片的文件名和路径
            img_filename = f"image_{img_idx}.png"
            img_idx += 1
            img_path = os.path.join(image_dir, img_filename)
            # 将图片数据写入文件
            with open(img_path, 'wb') as img_file:
                img_file.write(img_data)
            print(f'图片已保存到：{img_path}')
            fout.write(json.dumps({
                "keyword": keyword,
                "image": img_filename,
                "exp": jieshi[keyword],
            }, ensure_ascii=False)+"\n")

fout.close()
